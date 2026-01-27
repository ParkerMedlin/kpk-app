from __future__ import annotations

import logging
import os
from decimal import Decimal, InvalidOperation
from io import BytesIO
from threading import Lock
from typing import Dict, Optional, Tuple

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from zipfile import BadZipFile
import xlrd
from xlrd.biffh import XLRDError

LOGGER = logging.getLogger(__name__)

HEADER_ROW_INDEX = 3  # 1-based index, matches legacy sheet layout.
ITEM_HEADER = 'itemcode'
EST_LANDED_HEADER = 'est landed cost'
EST_LANDED_ALIASES = ['stdcost', 'std cost', 'standard cost', 'est landed cost', 'estlandedcost']
NEXT_COST_HEADER = 'next cost'
NEXT_COST_ALIASES = ['next cost', 'nextcost']
COST_FILE_PREFIX = 'zzz- master cost'
SUPPORTED_COST_EXTENSIONS = ('.xls', '.xlsx', '.xlsm')


class PurchasingCostParseError(Exception):
    """Raised when a workbook cannot be parsed into purchasing costs."""


_CACHE_LOCK = Lock()
_CACHE: Dict[str, object] = {
    'path': None,
    'mtime': None,
    'data': {},
}
_BOTH_CACHE: Dict[str, object] = {
    'path': None,
    'mtime': None,
    'data': {},
}


def _resolve_workbook_path(base_path: Optional[str]) -> Optional[str]:
    if not base_path:
        LOGGER.warning("COST_WORKBOOK_PATH is not configured; server workbook unavailable.")
        return None

    expanded = os.path.expanduser(base_path)
    if os.path.isdir(expanded):
        try:
            entries = os.listdir(expanded)
        except OSError as exc:
            LOGGER.warning(
                "Unable to list workbook directory '%s': %s", expanded, exc
            )
            return None

        matches = []
        prefix = COST_FILE_PREFIX.lower()
        for name in entries:
            lower = name.lower()
            if not lower.startswith(prefix):
                continue
            if not lower.endswith(tuple(ext for ext in SUPPORTED_COST_EXTENSIONS)):
                continue
            matches.append(os.path.join(expanded, name))

        if not matches:
            LOGGER.warning(
                "No workbook matching '%s*' with extensions %s found in '%s'.",
                COST_FILE_PREFIX,
                SUPPORTED_COST_EXTENSIONS,
                expanded,
            )
            return None

        matches.sort(key=lambda path: os.path.getmtime(path), reverse=True)
        resolved = matches[0]
        LOGGER.info("Using latest cost workbook '%s'.", resolved)
        return resolved

    if os.path.isfile(expanded):
        return expanded

    LOGGER.warning(
        "Configured COST_WORKBOOK_PATH '%s' does not exist.", expanded
    )
    return None


def load_default_purchasing_costs() -> Tuple[Dict[str, Dict[str, Decimal]], Optional[str]]:
    """
    Load (and cache) purchasing costs from the configured workbook path.

    Returns a tuple of (data, label) where data is an item-cost map and label is the basename
    of the workbook that supplied the data (or None if no workbook is configured).
    """
    configured_path = getattr(settings, 'COST_WORKBOOK_PATH', None)
    workbook_path = _resolve_workbook_path(configured_path)
    if not workbook_path:
        return {}, None

    try:
        stat = os.stat(workbook_path)
    except OSError as exc:
        LOGGER.warning("Unable to read cost workbook '%s': %s", workbook_path, exc)
        return {}, None

    with _CACHE_LOCK:
        if (
            _CACHE['path'] == workbook_path
            and _CACHE['mtime'] == stat.st_mtime
            and isinstance(_CACHE['data'], dict)
        ):
            return (
                {code: values.copy() for code, values in _CACHE['data'].items()},
                os.path.basename(workbook_path),
            )

    with open(workbook_path, 'rb') as workbook_file:
        data = workbook_file.read()

    parsed = _parse_costs_from_bytes(data, source_name=workbook_path)

    with _CACHE_LOCK:
        _CACHE['path'] = workbook_path
        _CACHE['mtime'] = stat.st_mtime
        _CACHE['data'] = parsed

    return (
        {code: values.copy() for code, values in parsed.items()},
        os.path.basename(workbook_path),
    )


def load_purchasing_costs_from_file(uploaded_file) -> Tuple[Dict[str, Dict[str, Decimal]], str]:
    """
    Parse purchasing costs from an uploaded workbook (InMemoryUploadedFile/File).

    Returns the parsed map and the original filename. Raises PurchasingCostParseError on failure.
    """
    try:
        data = uploaded_file.read()
    finally:
        try:
            uploaded_file.seek(0)
        except Exception:
            pass

    parsed = _parse_costs_from_bytes(data, source_name=getattr(uploaded_file, 'name', None))
    return parsed, uploaded_file.name


def _parse_costs_from_bytes(blob: bytes, source_name: Optional[str] = None) -> Dict[str, Dict[str, Decimal]]:
    if not blob:
        raise PurchasingCostParseError("The uploaded workbook is empty.")

    reason = source_name or 'the uploaded workbook'
    workbook_error = None
    xls_book = None
    try:
        workbook = load_workbook(filename=BytesIO(blob), read_only=True, data_only=True)
    except (InvalidFileException, BadZipFile, KeyError):
        workbook = None
    except Exception as exc:
        workbook = None
        workbook_error = exc

    if workbook:
        try:
            rows = _iter_openpyxl_rows(workbook.worksheets[0])
            return _parse_rows(rows)
        finally:
            workbook.close()

    try:
        xls_book = xlrd.open_workbook(file_contents=blob)
        sheet = xls_book.sheet_by_index(0)
        rows = _iter_xls_rows(sheet)
        return _parse_rows(rows)
    except XLRDError as exc:
        raise PurchasingCostParseError(
            f"Unable to read {reason}; please supply an .xlsx/.xlsm or legacy .xls workbook."
        ) from exc
    except Exception as exc:
        if workbook_error:
            raise PurchasingCostParseError("Unable to read the cost workbook.") from workbook_error
        raise PurchasingCostParseError("Unable to read the cost workbook.") from exc
    finally:
        if xls_book:
            try:
                xls_book.release_resources()
            except Exception:
                pass


def _iter_openpyxl_rows(worksheet):
    for row in worksheet.iter_rows(values_only=True):
        yield row


def _iter_xls_rows(sheet):
    for row_idx in range(sheet.nrows):
        yield sheet.row_values(row_idx)


def _parse_rows(row_iter) -> Dict[str, Dict[str, Decimal]]:
    header_row_index = HEADER_ROW_INDEX - 1
    headers = None
    required_columns = None
    result: Dict[str, Dict[str, Decimal]] = {}

    for idx, row in enumerate(row_iter):
        if idx < header_row_index:
            continue
        if idx == header_row_index:
            headers = _extract_headers_from_row(row)
            required_columns = _resolve_required_columns(headers)
            continue

        if not required_columns:
            raise PurchasingCostParseError("Missing header row in the cost workbook.")

        item_code_raw = _get_cell_value(row, required_columns.get(ITEM_HEADER))
        if not item_code_raw:
            continue

        # Normalize item code: strip whitespace and convert to string
        item_code = str(item_code_raw).strip() if item_code_raw else ''
        if not item_code:
            continue

        est_cost = _to_decimal_safe(_get_cell_value(row, required_columns.get(EST_LANDED_HEADER)))
        next_cost = _to_decimal_safe(_get_cell_value(row, required_columns.get(NEXT_COST_HEADER)))

        if est_cost > 0:
            result[item_code] = {'cost': est_cost, 'source': 'Est Landed'}
        elif next_cost > 0:
            result[item_code] = {'cost': next_cost, 'source': 'Next Cost'}

    return result


def _get_cell_value(row, index):
    if index is None:
        return None
    if row is None:
        return None
    if index >= len(row):
        return None
    return row[index]


def _extract_headers_from_row(row) -> Dict[int, str]:
    headers: Dict[int, str] = {}
    if not row:
        raise PurchasingCostParseError(
            f"Could not read headers from row {HEADER_ROW_INDEX} of the workbook."
        )
    for idx, cell in enumerate(row):
        value = _normalize_string(cell)
        if value:
            headers[idx] = value
    if not headers:
        raise PurchasingCostParseError(
            f"Could not read headers from row {HEADER_ROW_INDEX} of the workbook."
        )
    return headers


def _resolve_required_columns(headers: Dict[int, str]) -> Dict[str, int]:
    header_lookup = {name.lower(): index for index, name in headers.items()}

    item_col = header_lookup.get(ITEM_HEADER)
    if item_col is None:
        raise PurchasingCostParseError(f"The workbook is missing required column: {ITEM_HEADER.title()}.")

    est_landed_col = None
    for alias in EST_LANDED_ALIASES:
        if alias in header_lookup:
            est_landed_col = header_lookup[alias]
            break

    next_cost_col = None
    for alias in NEXT_COST_ALIASES:
        if alias in header_lookup:
            next_cost_col = header_lookup[alias]
            break

    if est_landed_col is None:
        raise PurchasingCostParseError(
            f"The workbook is missing a cost column. Expected one of: {', '.join(a.title() for a in EST_LANDED_ALIASES)}."
        )

    if next_cost_col is None:
        raise PurchasingCostParseError(
            f"The workbook is missing a next cost column. Expected one of: {', '.join(a.title() for a in NEXT_COST_ALIASES)}."
        )

    return {
        ITEM_HEADER: item_col,
        EST_LANDED_HEADER: est_landed_col,
        NEXT_COST_HEADER: next_cost_col,
    }


def _normalize_string(value) -> str:
    if value is None:
        return ''
    value = str(value).strip()
    return value.upper() if value else ''


def _to_decimal_safe(value) -> Decimal:
    if value is None or value == '':
        return Decimal('0')
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal('0')


def _parse_rows_with_both_costs(row_iter) -> Dict[str, Dict[str, Decimal]]:
    """
    Parse workbook rows to extract BOTH est_landed_cost and next_cost for cost impact analysis.

    Returns dict mapping item_code to {'est_landed': Decimal, 'next_cost': Decimal}
    """
    header_row_index = HEADER_ROW_INDEX - 1
    headers = None
    required_columns = None
    result: Dict[str, Dict[str, Decimal]] = {}

    for idx, row in enumerate(row_iter):
        if idx < header_row_index:
            continue
        if idx == header_row_index:
            headers = _extract_headers_from_row(row)
            required_columns = _resolve_required_columns(headers)
            continue

        if not required_columns:
            raise PurchasingCostParseError("Missing header row in the cost workbook.")

        item_code_raw = _get_cell_value(row, required_columns.get(ITEM_HEADER))
        if not item_code_raw:
            continue

        # Normalize item code: strip whitespace and convert to string
        item_code = str(item_code_raw).strip() if item_code_raw else ''
        if not item_code:
            continue

        est_cost = _to_decimal_safe(_get_cell_value(row, required_columns.get(EST_LANDED_HEADER)))
        next_cost = _to_decimal_safe(_get_cell_value(row, required_columns.get(NEXT_COST_HEADER)))

        result[item_code] = {
            'est_landed': est_cost,
            'next_cost': next_cost
        }

    return result


def load_purchasing_costs_with_both_values(uploaded_file=None) -> Tuple[Dict[str, Dict[str, Decimal]], str]:
    """
    Load purchasing costs with BOTH est_landed and next_cost for cost impact analysis.

    Args:
        uploaded_file: Optional uploaded file. If None, uses default server workbook.

    Returns:
        Tuple of (cost_map, workbook_name) where cost_map has both est_landed and next_cost
    """
    if uploaded_file:
        try:
            data = uploaded_file.read()
        finally:
            try:
                uploaded_file.seek(0)
            except Exception:
                pass
        parsed = _parse_costs_with_both_from_bytes(data, source_name=getattr(uploaded_file, 'name', None))
        return parsed, uploaded_file.name

    # Load from default server workbook (cached by path + mtime)
    configured_path = getattr(settings, 'COST_WORKBOOK_PATH', None)
    workbook_path = _resolve_workbook_path(configured_path)
    if not workbook_path:
        return {}, None

    try:
        stat = os.stat(workbook_path)
    except OSError as exc:
        LOGGER.warning("Unable to read cost workbook '%s': %s", workbook_path, exc)
        return {}, None

    with _CACHE_LOCK:
        if (
            _BOTH_CACHE['path'] == workbook_path
            and _BOTH_CACHE['mtime'] == stat.st_mtime
            and isinstance(_BOTH_CACHE['data'], dict)
        ):
            return _BOTH_CACHE['data'], os.path.basename(workbook_path)

    try:
        with open(workbook_path, 'rb') as workbook_file:
            data = workbook_file.read()
        parsed = _parse_costs_with_both_from_bytes(data, source_name=workbook_path)
    except Exception as exc:
        LOGGER.warning("Unable to load cost workbook with both values: %s", exc)
        return {}, None

    with _CACHE_LOCK:
        _BOTH_CACHE['path'] = workbook_path
        _BOTH_CACHE['mtime'] = stat.st_mtime
        _BOTH_CACHE['data'] = parsed

    return parsed, os.path.basename(workbook_path)


def _parse_costs_with_both_from_bytes(blob: bytes, source_name: Optional[str] = None) -> Dict[str, Dict[str, Decimal]]:
    """Parse workbook bytes and return both est_landed and next_cost values."""
    if not blob:
        raise PurchasingCostParseError("The uploaded workbook is empty.")

    reason = source_name or 'the uploaded workbook'
    workbook_error = None
    xls_book = None
    try:
        workbook = load_workbook(filename=BytesIO(blob), read_only=True, data_only=True)
    except (InvalidFileException, BadZipFile, KeyError):
        workbook = None
    except Exception as exc:
        workbook = None
        workbook_error = exc

    if workbook:
        try:
            rows = _iter_openpyxl_rows(workbook.worksheets[0])
            return _parse_rows_with_both_costs(rows)
        finally:
            workbook.close()

    try:
        xls_book = xlrd.open_workbook(file_contents=blob)
        sheet = xls_book.sheet_by_index(0)
        rows = _iter_xls_rows(sheet)
        return _parse_rows_with_both_costs(rows)
    except XLRDError as exc:
        raise PurchasingCostParseError(
            f"Unable to read {reason}; please supply an .xlsx/.xlsm or legacy .xls workbook."
        ) from exc
    except Exception as exc:
        if workbook_error:
            raise PurchasingCostParseError("Unable to read the cost workbook.") from workbook_error
        raise PurchasingCostParseError("Unable to read the cost workbook.") from exc
    finally:
        if xls_book:
            try:
                xls_book.release_resources()
            except Exception:
                pass
