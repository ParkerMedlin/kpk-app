import os
from openpyxl import load_workbook

def print_bad_component_item_codes():
    # Loop through the main folder and then the AF/ww folders, building list of all filepaths.
    file_list = []
    for root, dirs, files in os.walk(os.path.expanduser('~\\Desktop\\blendSheets')):
        for file in files:
            if not file.endswith('.db') and not file.endswith('.tmp'):
                file_list.append(os.path.join(root,file))
    for i, source_file_path in enumerate(file_list):
        try:
            if "~" in source_file_path:
                continue
            # Load the workbook
            wb = load_workbook(filename=source_file_path, read_only=True, keep_vba=True)

            # Select a sheet
            this_worksheet = wb['BlendSheet']
            item_code_value = this_worksheet.cell(row=1, column=9).value
            # Get the formula from a cell
            cell = this_worksheet['C5']
            if not isinstance(cell.value, str):
                print(source_file_path)
                continue
            if not cell.value.startswith('='):
                print(source_file_path)

            # Create a list to store the values
            column_a_values = []

            # Loop through rows 1 to 26 in column D
            for row in range(5, 20):
                cell_value = this_worksheet.cell(row=row, column=1).value
                column_a_values.append(cell_value)

            # Remove empty cells from the list
            unwanted_values = ['Manufacturing Record','Product:','PROCESS PREPARATION',
                               'Equipment to come in contact with product will be cleaned',
                               'BATCH PREPARATION','StepNumber','ItemCode','VW','Prepared By:',
                               'WATER','NEW']
            column_a_values = [ value for value in column_a_values if value is not None ]
            column_a_values = [ value for value in column_a_values if str(value) not in unwanted_values ]
            column_a_values = [ value for value in column_a_values if '*****' not in str(value) and '^^^^^' not in str(value) and '=' not in str(value) ]
            
            for value in column_a_values:
                if len(str(value)) < 6:
                    print(item_code_value, value)

        except Exception as e:
            print(source_file_path)
            print(str(e))
            continue

print_bad_component_item_codes()
