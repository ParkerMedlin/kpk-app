import pyexcel as pe
import pandas as pd
import os
import csv
from openpyxl import load_workbook
import shutil
import datetime as dt

def get_blend_procedures():
    # Loop through the main folder and then the AF/ww folders, building list of all filepaths.
    file_list = []
    for root, dirs, files in os.walk(os.path.expanduser('~\\Desktop\\blendSheets')):
        for file in files:
            if not file.endswith('.db') and not file.endswith('.tmp'):
                file_list.append(os.path.join(root,file))

    data_frames = []
    # For each file, create a dataframe and then append that dataframe to the csv.
    for i, source_file_path in enumerate(file_list):
        try:
            if "~" in source_file_path:
                continue
            if not source_file_path.endswith('.xlsx'):
                continue
            this_workbook = load_workbook(source_file_path, data_only=True)
            this_worksheet = this_workbook['BlendSheet']
            item_code_value = this_worksheet.cell(row=1, column=9).value
            formula_reference_number = this_worksheet.cell(row=3, column=1).value
            product_density = this_worksheet.cell(row=3, column=10).value
            datetime = dt.datetime.now()

            # formula_reference_number
            # product_name
            # blend_number
            # product_density
            # component_item_code
            # percent_weight_of_total
            # sequence
            # blend_instructions
            # date

            # create the dataframe for this blendsheet.
            instruction_set = pd.read_excel(source_file_path, 'BlendSheet', skiprows = 5, usecols = 'A:B,E,F', dtype=object)

            instruction_set['blend_item_code'] = str(item_code_value)
            table_name = 'blend_formula_component'

            data_frames.append(instruction_set)

            #write to csv
            # instruction_set.to_csv(os.path.expanduser(r'~\\Desktop\\blendinstructions.csv'), mode='a', header=False, index=False) # Write to the csv in our folder

        except Exception as e:
            print(source_file_path)
            print(str(e))
            continue
    
    final_df = pd.concat(data_frames).iloc[:, :-2]
    final_df.columns = ["step_number","step_description","unit","component_item_code","blend_item_code" ]
    print(final_df)
    final_df.to_csv(os.path.expanduser('~\\Desktop')+"\\blendinstructions.csv", index=False)
    # final_df.to_csv(os.path.expanduser('~\\Documents')+"\\kpk-app\\db_imports\\blendinstructions.csv", index=False)

get_blend_procedures()