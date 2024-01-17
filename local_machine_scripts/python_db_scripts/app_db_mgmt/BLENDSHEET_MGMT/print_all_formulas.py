import os
from openpyxl import load_workbook
import pandas as pd
import psycopg2

def print_all_formulas():
    connection_postgres = psycopg2.connect('postgresql://postgres:blend2021@localhost:5432/blendversedb')
    cursor_postgres = connection_postgres.cursor()

    cursor_postgres.execute("""
        select itemcode, shipweight from ci_item where itemcodedesc not like '/%'
    """)
    # Fetch all rows
    rows = cursor_postgres.fetchall()

    # Create a dictionary from the rows
    itemcode_shipweight_dict = {row[0]: row[1] for row in rows}

    connection_postgres.commit()
    cursor_postgres.close()
    connection_postgres.close()

    # Loop through the main folder and then the AF/ww folders, building list of all filepaths.
    file_list = []
    for root, dirs, files in os.walk(os.path.expanduser('~\\Desktop\\blendSheets')):
        for file in files:
            if not file.endswith('.db') and not file.endswith('.tmp'):
                file_list.append(os.path.join(root,file))

    formulas_list = []
    for i, source_file_path in enumerate(file_list):
        try:
            if "~" in source_file_path:
                continue
            # Load the workbook
            wb = load_workbook(filename=source_file_path, read_only=True, keep_vba=True)

            # Select a sheet
            this_worksheet = wb['BlendSheet']
            item_code_value = this_worksheet.cell(row=1, column=9).value
            
             # Look for 'WATER' in the first 25 rows of column A
            for row in range(1, 26):
                this_item_code_cell = str(this_worksheet.cell(row=row, column=1).value)
                skip_values = ["StarBrite","Formula Reference No.","WATER","030143","Prepared By:","VW","ItemCode"]

                if this_item_code_cell not in skip_values:
                    if this_worksheet.cell(row=row, column=4).value:
                        quantity_cell_formula = str(this_worksheet.cell(row=row, column=4).value.replace('=', ''))
                        weightmatch = 'Match' if str(itemcode_shipweight_dict.get(this_item_code_cell, 0)) in quantity_cell_formula else ''
                        if '/' in quantity_cell_formula:
                            # Add to list instead of printing
                            formulas_list.append({'BLEND': item_code_value, 
                                                  'Itemcode': this_item_code_cell,
                                                  'Formula': quantity_cell_formula, 
                                                  'Filepath' : source_file_path,
                                                  'Shipweight' : str(itemcode_shipweight_dict.get(this_item_code_cell, 0)),
                                                  'Weightmatch' : weightmatch
                                                  })

        except Exception as e:
            print(source_file_path)
            print(str(e))

    df = pd.DataFrame(formulas_list)
    
    df.to_csv(os.path.expanduser('~\\Desktop\\blendSheetFormulas.csv'), index=False)
    print(df)

print_all_formulas()