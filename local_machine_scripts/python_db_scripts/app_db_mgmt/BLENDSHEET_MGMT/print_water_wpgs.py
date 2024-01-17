import os
from openpyxl import load_workbook
import pandas as pd

def print_water_formulas():
    # Loop through the main folder and then the AF/ww folders, building list of all filepaths.
    file_list = []
    for root, dirs, files in os.walk(os.path.expanduser('~\\Desktop\\blendSheets')):
        for file in files:
            if not file.endswith('.db') and not file.endswith('.tmp'):
                file_list.append(os.path.join(root,file))

    water_calcs_list = []
    for i, source_file_path in enumerate(file_list):
        try:
            if "~" in source_file_path:
                continue
            # Load the workbook
            wb = load_workbook(filename=source_file_path, read_only=True, keep_vba=True)

            # Select a sheet
            this_worksheet = wb['BlendSheet']
            item_code_value = this_worksheet.cell(row=1, column=9).value
            
             # Look for 'WATER' in the first 25 rows of column A
            for row in range(1, 26):
                this_item_code_cell = this_worksheet.cell(row=row, column=1).value
                possible_water_values = ['WATER', '030143', "'030143", 30143]
                if this_item_code_cell in possible_water_values:
                    offset_cell_value = str(this_worksheet.cell(row=row, column=4).value.replace('=', ''))
                    if '/' in offset_cell_value:
                        # Add to list instead of printing
                        if '8.34' in offset_cell_value:
                            water_calcs_list.append({'BLEND': item_code_value, 'Water Type': this_item_code_cell, 'Formula': offset_cell_value, 'Filepath' : source_file_path})
                        else:
                            water_calcs_list.insert(0, {'BLEND': item_code_value, 'Water Type': this_item_code_cell, 'Formula': offset_cell_value, 'Filepath' : source_file_path})
                        print(source_file_path)

        except Exception as e:
            print(source_file_path)
            print(str(e))
            continue

    df = pd.DataFrame(water_calcs_list)
    df.to_csv(os.path.expanduser('~\\Desktop\\water_calcs.csv'), index=False)
    print(df)

print_water_formulas()