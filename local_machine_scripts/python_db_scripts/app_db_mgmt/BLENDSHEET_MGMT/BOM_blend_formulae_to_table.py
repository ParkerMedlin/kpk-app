import pyexcel as pe
import pandas as pd
import os
import csv
from openpyxl import load_workbook
import shutil
import datetime as dt

def get_blend_ingredients():
    file_list = []
    for root, dirs, files in os.walk(os.path.expanduser('~\\Desktop\\testers')):
        for file in files:
            if not file.endswith('.db') and not file.endswith('.tmp'):
                file_list.append(os.path.join(root,file))

    data_frames = []
    for i, source_file_path in enumerate(file_list):
        try:
            if "~" in source_file_path or not source_file_path.endswith('.xlsx'):
                continue
            this_workbook = load_workbook(source_file_path, data_only=True)
            this_worksheet = this_workbook.worksheets[0]

            # Assuming the rows you're interested in start from row 6 to 20 (as per skiprows=5 and nrows=15)
            # and column D contains the calculated values you're interested in.
            calculated_values = []
            for row in range(6, 21):  # Adjust the range as necessary
                cell_value = this_worksheet.cell(row=row, column=4).value  # Column 4 corresponds to column D
                calculated_values.append(cell_value)

            # Now, you can create a DataFrame from the calculated values
            # Assuming you want to pair these with the component item codes from column A
            component_item_codes = [this_worksheet.cell(row=row, column=1).value for row in range(6, 21)]
            ingredient_set = pd.DataFrame({
                'component_item_code': component_item_codes,
                'calculated_value': calculated_values
            })
            ingredient_set.dropna(how='all', inplace=True)  # Removes rows where all elements are NaN

            ingredient_set['blend_item_code'] = str(this_worksheet.cell(row=1, column=9).value)
            data_frames.append(ingredient_set)

        except Exception as e:
            print(source_file_path)
            print(str(e))
            continue
    
    final_df = pd.concat(data_frames)
    final_df.columns = ["blend_item_code", "component_item_code", "amount"]
    print(final_df)
    final_df.to_excel(os.path.expanduser('~\\Desktop')+"\\blendingredients2.xlsx", index=False)

get_blend_ingredients()