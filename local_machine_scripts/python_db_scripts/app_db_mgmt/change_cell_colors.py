import os
import time
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.formatting.rule import Rule
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.styles.differential import DifferentialStyle
from PIL import Image
from openpyxl_image_loader import SheetImageLoader



# def convert_excel_to_jpg(worksheet, file_path):
#     #calling the image_loader
#     image_loader = SheetImageLoader(worksheet)
#     image = image_loader.get('A1:K52')
#     image.save(f'{file_path}.jpg')

# import excel2img
# import win32com.client
# from PIL import ImageGrab

# def export_blendsheet_to_jpg_v2(file_path):
#     o = win32com.client.Dispatch('Excel.Application')
#     o.visible = False

#     wb = o.Workbooks.Open(file_path)
#     ws = wb.Sheet('BlendSheet')

#     ws.Range(ws.Cells(1,1),ws.Cells(10,52)).Copy()
#     img_root = "U:\\qclab\\My Documents\\Lab Sheets 04 10 07\\Blend Sheets 06 15 07\\BLEND SHEETS 06 15 07\\result_images"
#     img_path = os.path.join(img_root, f"{file_name}.jpg")
#     img = ImageGrab.grabclipboard()
#     imgFile = img_path
#     img.save(imgFile)


# def export_blendsheet_to_jpg(file_path, file_name):
#     img_root = "U:\\qclab\\My Documents\\Lab Sheets 04 10 07\\Blend Sheets 06 15 07\\BLEND SHEETS 06 15 07\\result_images"
#     img_path = os.path.join(img_root, f"{file_name}.png")
#     excel2img.export_img(file_path,img_path,"BlendSheet!A1:K52")

def update_theory_gal_cell(worksheet):
    worksheet['F3'] = 1000

def set_water_code(worksheet, file_name):
    for cell in worksheet['B']:
        if 'add water.' in str(cell.value).lower() or 'charge vessel with water.' in str(cell.value).lower():
            worksheet.cell(row=cell.row, column=6).value = 'WATER'
            print(f'WATER added to cell{cell.coordinate} in {file_name}')

def change_qty_cells_color(worksheet):
    gal_fill_pattern = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
    lbs_fill_pattern = PatternFill(start_color='fcd87c', end_color='fcd87c', fill_type='solid')
    grams_fill_pattern = PatternFill(start_color='e57cfc', end_color='e57cfc', fill_type='solid')
    gram_font = Font(name='Calibri', size=11, italic=True)
    gram_alignment = Alignment(horizontal='center')
    lbs_alignment = Alignment(horizontal='right')

    for cell in worksheet['E']:
        if 'gal' in str(cell.value).lower():
            cell.fill = gal_fill_pattern
            a_value = str(worksheet.cell(row=cell.row, column=1).value)
            for lower_cell in worksheet['F']:
                if a_value in str(lower_cell.value):
                    left_col = get_column_letter(lower_cell.column - 1)
                    left_cell = left_col + str(lower_cell.row)
                    worksheet[left_cell].fill = gal_fill_pattern
        if 'lbs' in str(cell.value).lower() or 'lb' in str(cell.value).lower():
            cell.fill = lbs_fill_pattern
            cell.alignment = lbs_alignment
            a_value = str(worksheet.cell(row=cell.row, column=1).value)
            for lower_cell in worksheet['F']:
                if a_value in str(lower_cell.value):
                    left_col = get_column_letter(lower_cell.column - 1)
                    left_cell = left_col + str(lower_cell.row)
                    worksheet[left_cell].fill = lbs_fill_pattern
                    worksheet[left_cell].alignment = lbs_alignment
        if 'gram' in str(cell.value).lower() or 'grams' in str(cell.value).lower():
            cell.fill = grams_fill_pattern
            cell.font = gram_font
            cell.alignment = gram_alignment
            a_value = str(worksheet.cell(row=cell.row, column=1).value)
            for lower_cell in worksheet['F']:
                if a_value in str(lower_cell.value):
                    left_col = get_column_letter(lower_cell.column - 1)
                    left_cell = left_col + str(lower_cell.row)
                    worksheet[left_cell].fill = grams_fill_pattern
                    worksheet[left_cell].font = gram_font
                    worksheet[left_cell].alignment = gram_alignment

            # left_col = get_column_letter(cell.column - 1)
            # left_cell = left_col + str(cell.row)
            # worksheet[left_cell].fill = fill_pattern # Apply the fill pattern to the left cell

folder_paths = [
    'U:\\qclab\\My Documents\\Lab Sheets 04 10 07\\Blend Sheets 06 15 07\\BLEND SHEETS 06 15 07',
    # "C:/Users/pmedlin/Desktop/testing",
    # "C:/Users/pmedlin/Desktop/testing/testmoar",
    "U:\\qclab\\My Documents\\Lab Sheets 04 10 07\\Blend Sheets 06 15 07\\BLEND SHEETS 06 15 07\\1) -50 RVAF",
    "U:\\qclab\\My Documents\\Lab Sheets 04 10 07\\Blend Sheets 06 15 07\\BLEND SHEETS 06 15 07\\2) -60 RVAF",
    "U:\\qclab\\My Documents\\Lab Sheets 04 10 07\\Blend Sheets 06 15 07\\BLEND SHEETS 06 15 07\\3) -100RVAF",
    "U:\\qclab\\My Documents\\Lab Sheets 04 10 07\\Blend Sheets 06 15 07\\BLEND SHEETS 06 15 07\\4) -200RVAF",
    "U:\\qclab\\My Documents\\Lab Sheets 04 10 07\\Blend Sheets 06 15 07\\BLEND SHEETS 06 15 07\\5) -SPLASH W-W\\Drying Agent Premix",
    "U:\\qclab\\My Documents\\Lab Sheets 04 10 07\\Blend Sheets 06 15 07\\BLEND SHEETS 06 15 07\\5) -SPLASH W-W\\NON-repel formulas",
    "U:\\qclab\\My Documents\\Lab Sheets 04 10 07\\Blend Sheets 06 15 07\\BLEND SHEETS 06 15 07\\5) -SPLASH W-W\\REPEL formulas"
    ]

# bw = '90400.B (VIOLET Boat Wash 6-36AP).xlsx'
# bw2 = '602000_(Boat Wash 3-135A).xlsx'

file_count = 0
for folder_path in folder_paths:
    for file_name in os.listdir(folder_path):
        if file_name.endswith(".xlsx") and '~' not in file_name: #check if the file is an xlsx file
            try:
                file_path = os.path.join(folder_path, file_name)
                workbook = load_workbook(file_path)
                worksheet = workbook["BlendSheet"]
                # convert_excel_to_jpg(worksheet, file_name)
                worksheet.protection.disable()
                set_water_code(worksheet, file_name)
                change_qty_cells_color(worksheet)
                update_theory_gal_cell(worksheet)
                worksheet.protection.enable()
                file_count += 1
                workbook.save(file_path)
            except Exception as e:
                print(f'error {str(e)} with {file_path}')
                continue
print(str(file_count) + "files found")




# def update_conditional_formatting(worksheet):
#     gal_style = DifferentialStyle(fill=PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid'))
#     lb_style = DifferentialStyle(fill=PatternFill(start_color='fcd87c', end_color='fcd87c', fill_type='solid'), alignment=Alignment(horizontal='right'))
#     gram_style = DifferentialStyle(font=Font(italic=True), fill=PatternFill(start_color='e57cfc', end_color='e57cfc', fill_type='solid'), alignment=Alignment(horizontal='center'))

#     # Define the formatting rules
#     gal_rule = Rule(type='cellIs', dxf=gal_style, formula=['=E1="gal"'])
#     lb_rule = Rule(type='cellIs', dxf=lb_style, formula=['=OR(E1="lbs", "lb"'])
#     gram_rule = Rule(type='cellIs', dxf=gram_style, formula=['=OR(E1="g", E1="gram", E1="grams")'])

#     # Apply the rules to column E
#     for row in range(1, worksheet.max_row + 1):
#         cell = worksheet.cell(row=row, column=5)
#         worksheet.conditional_formatting.add(cell.coordinate, gal_rule)
#         worksheet.conditional_formatting.add(cell.coordinate, lb_rule)
#         worksheet.conditional_formatting.add(cell.coordinate, gram_rule)
